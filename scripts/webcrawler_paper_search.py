import os
import requests
from bs4 import BeautifulSoup, NavigableString
import re
from random import choice
import time
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import PyPDF2
from nltk.tokenize import sent_tokenize
import requests
import cloudscraper
import random
from scholarly import scholarly, ProxyGenerator
import csv
import traceback
from openpyxl.comments import Comment
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import cloudscraper
import nltk
from nltk import ne_chunk, pos_tag, word_tokenize
from nltk.tree import Tree
import pubchempy as pcp
import unicodedata
import subprocess
from itertools import permutations
from urllib.parse import urlencode
import backoff
from urllib3.util.retry import Retry
nltk.download('punkt', quiet=True)
nltk.download('averaged_perceptron_tagger', quiet=True)
nltk.download('maxent_ne_chunker', quiet=True)
nltk.download('words', quiet=True)




# Author information
AUTHOR_INFO = """
###########################################################################
╔════════════════════════════════════════════════════════════════════╗
║                                                                    ║
║                    Scholarly_Paper_Crawler                         ║
║                                                                    ║
╠════════════════════════════════════════════════════════════════════╣
║  Author:  Suk Jin Mun                                              ║
║  Version: 1.0.0                                                    ║
║  Year:    2024                                                     ║
║  License: MIT                                                      ║
║  Email:   rOysJmUN[at]gMail.com (all lower case letters)           ║
╚════════════════════════════════════════════════════════════════════╝

This is an automated tool designed to streamline the process of
searching, retrieving, and analyzing academic papers from Google Scholar.
It automates the collection of research data, including paper details,
PDF downloads, and compound identification from abstracts.

For more information or to report issues, please contact the author.
###########################################################################
"""


USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 11.5; rv:90.0) Gecko/20100101 Firefox/90.0',
]

SCI_HUB_DOMAINS = [
    "https://sci-hub.se/",
    "https://sci-hub.wf/",
    "https://sci-hub.ee/",
    "https://sci-hub.re/",
    "https://sci-hub.ru/",
    "https://sci-hub.mksa.top/",
    "https://sci-hub.shop/",
    "https://scihub.wikicn.top/",
    "https://sci-hub.ren/"
]

pg = ProxyGenerator()
pg.FreeProxies()
scholarly.use_proxy(pg)
scholarly.set_timeout(10)  # Set a timeout for scholarly requests

# read_inputs function:
def read_inputs(file_path):
    inputs = {}
    with open(file_path, 'r') as file:
        lines = file.readlines()
        current_key = None
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
                
            if '>' in line:
                # Extract the number before the > as the key
                key_match = re.match(r'(\d+)>', line)
                if key_match:
                    current_key = key_match.group(1)
            elif current_key:
                # Store the value with its corresponding key
                inputs[current_key] = line
                current_key = None
                
    if '0' not in inputs:
        raise ValueError("API key not found in inputs.txt. Please add '0> Input API Key' section.")
    return inputs

def create_directories(base_dir, search_name):
    pdf_dir = os.path.join(base_dir, 'pdf_files', search_name)
    excel_dir = os.path.join(base_dir, 'csv_files')
    sentences_dir = os.path.join(base_dir, 'pdf_first_100_sentences', search_name)
    
    os.makedirs(pdf_dir, exist_ok=True)
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(sentences_dir, exist_ok=True)
    
    return pdf_dir, excel_dir, sentences_dir

def save_first_100_sentences(sentences, pdf_filename, sentences_dir):
    # Save first 100 sentences to a text file
    if not sentences or sentences.startswith("Failed to") or sentences.startswith("Error"):
        print(f"No valid sentences to save for {pdf_filename}")
        return
        
    # Generate the text filename from PDF filename
    text_filename = pdf_filename.replace('.pdf', '_first100.txt')
    text_filepath = os.path.join(sentences_dir, text_filename)
    
    try:
        with open(text_filepath, 'w', encoding='utf-8') as f:
            f.write(sentences)
        print(f"Saved first 100 sentences to: {text_filepath}")
    except Exception as e:
        print(f"Error saving sentences to file: {str(e)}")    

def extract_keywords(text, num_keywords=10):
    common_words = set(['the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by'])
    words = [word.lower() for word in re.findall(r'\b\w+\b', text) if word.lower() not in common_words]
    
    word_freq = {}
    for word in words:
        if len(word) > 2:
            word_freq[word] = word_freq.get(word, 0) + 1
    
    sorted_words = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)
    return [word for word, _ in sorted_words[:num_keywords]]

def extract_compound_name(text):
    # Ignore spin values (S = x or S=x)
    text = re.sub(r'\bS\s*=\s*\d+(/\d+)?', '', text)
    
    # Use regex to find potential chemical formulas with subscripts, allowing spaces between formula components
    chemical_formulas = re.findall(r'\b(?:[A-Z][a-z]?(?:\d+|₀-₉)*)+(?:\([^()]+\)(?:\d+|₀-₉)*)*\b', text)
    
    # Normalize chemical formulas by handling extra spaces between components (e.g., (CD3)4 NMn Cl3 -> (CD3)4NMnCl3)
    text = normalize_whitespace(text)

    # Filter to exclude cases where only one capital letter is present or there are fewer capital letters (e.g., "Non-linear", "Magnetic")
    potential_compounds = re.findall(r'\b[A-Z][a-z]*\s*(?:[A-Z][a-z]*\s*)*[A-Z][a-z]+\b', text)
    potential_compounds = [compound for compound in potential_compounds if len([c for c in compound if c.isupper()]) > 1]

    all_compounds = chemical_formulas + potential_compounds

    if all_compounds:
        # Try to get more information about the first compound using PubChem
        for compound in all_compounds:
            try:
                pcp_compounds = pcp.get_compounds(compound, 'name')
                if pcp_compounds:
                    formula = pcp_compounds[0].molecular_formula
                    if formula:
                        return formula
            except:
                pass
        return max(all_compounds, key=lambda x: len(x))
    
    return "No compound information found from abstract"

def generate_keyword_combinations(keywords):
    # Generate all possible combinations of keywords including individual keywords
    # and their permutations
    # Split and clean keywords
    keyword_list = [k.strip() for k in keywords.split(',')]
    
    # Generate all possible combinations
    all_combinations = []
    
    # Add individual keywords
    all_combinations.extend(keyword_list)
    
    # Generate permutations for lengths 2 to n
    for length in range(2, len(keyword_list) + 1):
        perms = list(permutations(keyword_list, length))
        for perm in perms:
            # Join with spaces to create search term
            all_combinations.append(' '.join(perm))
    
    return all_combinations

class ScholarScraper:
    def __init__(self, api_key):
        self.api_key = api_key
        self.session = self._create_session()
        self.base_url = "http://api.scraperapi.com"
        self.last_request_time = 0
        
    def _create_session(self):
        session = requests.Session()
        retries = Retry(
            total=5,
            backoff_factor=30,
            status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retries)
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        return session

    def _enforce_rate_limit(self):
        """Ensures minimum delay between requests"""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        if time_since_last < 20:
            time.sleep(20 - time_since_last)
        self.last_request_time = time.time()

    @backoff.on_exception(
        backoff.expo,
        (requests.exceptions.RequestException, Exception),
        max_tries=3
    )
    def search_google_scholar(self, query, start=0):
        self._enforce_rate_limit()
        
        # Fixed URL construction
        scholar_params = {
            "q": query,
            "hl": "en",
            "as_sdt": "0,5",
            "start": str(start)
        }
        
        encoded_url = f"https://scholar.google.com/scholar?{urlencode(scholar_params)}"
        
        params = {
            'api_key': self.api_key,
            'url': encoded_url,
            'render': 'true'
        }

        try:
            response = self.session.get(
                self.base_url,
                params=params,
                timeout=60
            )
            response.raise_for_status()
            return {'html': response.text, 'status': response.status_code}
        except Exception as e:
            print(f"Error in API request: {str(e)}")
            return None

def search_papers(keywords, max_results_per_combo, search_name, html_parsing_folder):
    # Generate keyword combinations
    keyword_list = generate_keyword_combinations(keywords)
    all_results = []

    # Create html_parsing folder if it doesn't exist
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    html_parsing_folder = os.path.join(root_dir, 'html_parsing', search_name)
    os.makedirs(html_parsing_folder, exist_ok=True)

    # Initialize scraper with your API key
    inputs_file = os.path.join(root_dir, 'inputs.txt')
    inputs = read_inputs(inputs_file)
    api_key = inputs.get('0')
    
    # Initialize scraper with API key from inputs
    scraper = ScholarScraper(api_key=api_key)

    print(f"\nWill search for {len(keyword_list)} keyword combinations:")
    for i, combo in enumerate(keyword_list, 1):
        print(f"{i}. {combo}")
    print()

    for keyword_combo in keyword_list:
        start = 0
        combo_results = 0  # Track results for current combination
        retry_count = 0
        max_retries = 3
        
        while combo_results < max_results_per_combo and retry_count < max_retries:
            try:
                # Use the scraper to get results
                result = scraper.search_google_scholar(keyword_combo, start)
                
                if not result:
                    print(f"Failed to get results for keyword: {keyword_combo}")
                    retry_count += 1
                    time.sleep(random.uniform(30, 60))
                    continue

                soup = BeautifulSoup(result['html'], 'html.parser')
                items = soup.select('.gs_r.gs_or.gs_scl')
                
                if not items:
                    print(f"No more results for keyword: {keyword_combo}")
                    break

                for index, item in enumerate(items):
                    if combo_results >= max_results_per_combo:
                        print(f"Reached the maximum number of results ({max_results_per_combo}) for: {keyword_combo}")
                        break

                    # Save raw HTML content
                    file_name = f"paper_{len(all_results) + 1}_raw.html"
                    file_path = os.path.join(html_parsing_folder, file_name)
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(str(item))

                    title_elem = item.select_one('.gs_rt')
                    authors_elem = item.select_one('.gs_a')
                    snippet_elem = item.select_one('.gs_rs')

                    title = extract_title(title_elem) if title_elem else "Unknown Title"
                    
                    # Check for duplicates
                    if any(paper['Title'] == title for paper in all_results):
                        print(f"Skipping duplicate paper: {title}")
                        continue

                    authors_year = authors_elem.text if authors_elem else "Unknown Authors, Unknown Year"
                    snippet = snippet_elem.text if snippet_elem else ""

                    authors = authors_year.split('-')[0] if '-' in authors_year else authors_year
                    year = re.search(r'\b\d{4}\b', authors_year)
                    year = year.group(0) if year else "Unknown"

                    extracted_keywords = ', '.join(extract_keywords(snippet))

                    scholar_link = item.select_one('.gs_rt a')['href'] if item.select_one('.gs_rt a') else "No link"

                    compound_name = extract_compound_name(title + " " + snippet)

                    result = {
                        'Title': title,
                        'Authors': authors,
                        'Year': year,
                        'Keywords': extracted_keywords,
                        'Google Scholar Link': scholar_link,
                        'Compound': compound_name,
                        'Search Terms': keyword_combo
                    }

                    all_results.append(result)
                    combo_results += 1
                    print(f"Processed paper {len(all_results)} (combination paper {combo_results}): {title}", flush=True)
                    print(f"Using search terms: {keyword_combo}", flush=True)
                    print(f"Snippet: {snippet}", flush=True)
                    print(f"Raw HTML saved to: {file_path}", flush=True)
                    print()

                if combo_results < max_results_per_combo:
                    next_button = soup.select_one('.gs_ico_nav_next')
                    if next_button and next_button.find_parent('a'):
                        print("Next page exists, moving to the next page.")
                        start += 10
                        print(f"Start value updated to: {start}")
                        time.sleep(random.uniform(60, 90))
                    else:
                        print(f"No more pages available for keyword: {keyword_combo}")
                        break

            except Exception as e:
                print(f"Error during search for '{keyword_combo}': {str(e)}")
                retry_count += 1
                time.sleep(random.uniform(60, 120))
                continue

        print(f"Completed search for combination: {keyword_combo}")
        print(f"Found {combo_results} papers for this combination")
        print(f"Total papers so far: {len(all_results)}\n")
        
        if keyword_combo != keyword_list[-1]:  # If not the last combination
            wait_time = random.uniform(60, 90)
            print(f"Waiting {wait_time:.2f} seconds before next combination...")
            time.sleep(wait_time)

    return all_results

def extract_svg_content(svg_elem):
    content = []
    aria_label = svg_elem.get('aria-label')
    if aria_label:
        return aria_label

    for child in svg_elem.descendants:
        if isinstance(child, NavigableString):
            content.append(child.strip())
        elif child.name == 'path' and child.get('aria-label'):
            # Handle known common fractions (like 1/2)
            if 'S=12' in child['aria-label']:  # Example case for fractions
                content.append("S = 1/2")
            else:
                content.append(child['aria-label'])
    return ' '.join(content)

def handle_subscripts_superscripts(text):
    def to_sub_or_sup(match, script_type):
        return ''.join(unicodedata.lookup(f"{script_type} {c.upper()}") if c.isalnum() else c for c in match.group(1))

    text = re.sub(r'<sub>(.*?)</sub>', lambda m: to_sub_or_sup(m, "SUBSCRIPT"), text)
    text = re.sub(r'<sup>(.*?)</sup>', lambda m: to_sub_or_sup(m, "SUPERSCRIPT"), text)
    return text

def normalize_whitespace(text):
    # Fix spaces inside chemical formulas (e.g., (C D3) -> (CD3), K Cu F3 -> KCuF3)
    
    # Remove spaces between elements and numbers inside parentheses (e.g., (C D3)4 -> (CD3)4)
    text = re.sub(r'\(\s*([A-Za-z0-9]+)\s*\)', r'(\1)', text)
    
    # Remove spaces between elements and numbers inside parentheses (e.g., (C D3)4 -> (CD3)4)
    text = re.sub(r'\(\s*([A-Za-z0-9]+)\s*\)', r'(\1)', text)

    # Handle chemical formulas without spaces between elements and numbers (e.g., K Cu F3 -> KCuF3, Ca Cu2 O3 -> CaCu2O3)
    text = re.sub(r'\b([A-Z][a-z]?)(\d*)\s*([A-Z][a-z]?)(\d*)\s*([A-Z][a-z]?)(\d*)?\b', r'\1\2\3\4\5\6', text)

    # Ensure correct handling of multi-element formulas such as "(CD3)4 NMn Cl3" -> "(CD3)4NMnCl3"
    text = re.sub(r'\b([A-Z][a-z]?)(\d*)\s*([A-Z][a-z]?)\s*([A-Z][a-z]?)(\d*)\b', r'\1\2\3\4\5', text)

    # Handle numbers with elements (e.g., Cu 2 -> Cu2)
    text = re.sub(r'\b([A-Z][a-z]*)(\d*)\s+([A-Z][a-z]*)(\d*)\b', r'\1\2\3\4', text)
    
    # General cleanup for extra spaces
    text = re.sub(r'\s+', ' ', text).strip()

    return text

def clean_html(text):
    # Replace HTML tags with appropriate text representations
    text = re.sub(r'<i>(.*?)</i>', r'_\1_', text)  # Italics
    text = re.sub(r'<b>(.*?)</b>', r'**\1**', text)  # Bold
    text = re.sub(r'<[^>]+>', '', text)  # Remove any remaining tags
    return text

def parse_with_context(elem):
    if elem.name == 'span' and 'gs_fsvg' in elem.get('class', []):
        # This is likely a special symbol or formula
        return extract_svg_content(elem)
    elif elem.name in ['sub', 'sup']:
        return handle_subscripts_superscripts(elem.text)
    else:
        return elem.text

def extract_title(title_elem):
    if not title_elem:
        return "Unknown Title"

    try:
        full_title = []
        previous_content = None  # Track previous content to avoid duplicates
        
        for elem in title_elem.descendants:
            if isinstance(elem, NavigableString):
                content = elem.strip()
            elif elem.name == 'svg':
                svg_content = extract_svg_content(elem)
                content = svg_content.strip()
            elif elem.name in ['sub', 'sup']:
                content = handle_subscripts_superscripts(str(elem))
            elif elem.name == 'span' and 'gs_fsvg' in elem.get('class', []):
                svg_content = extract_svg_content(elem)
                content = svg_content.strip()
            elif elem.name == 'b':  # Processing bold text
                content = elem.text.strip()
            else:
                content = None  # Ignore other tags for now

            # Only append if content is not a duplicate of the previous content
            if content and content != previous_content:
                full_title.append(content)
                previous_content = content

        full_title = ' '.join(full_title)

        # Replace common HTML entities
        full_title = full_title.replace('&nbsp;', ' ')
        full_title = full_title.replace('&amp;', '&')

        full_title = handle_subscripts_superscripts(full_title)

        # Handle chemical formulas
        full_title = re.sub(r'([A-Z][a-z]*)(\d+)', r'\1\2', full_title)  # Keep chemical formulas like CaCu2O3 without subscripts
    
        # Remove any remaining HTML tags
        full_title = clean_html(full_title)

        # Handle special cases with two numbers as fractions (generalized)
        full_title = re.sub(r'S=(\d+)(\d+)', lambda m: f"S={m.group(1)}/{m.group(2)}", full_title)  # Specifically for spin values S=1/2

        # Handle cases where words are incorrectly joined
        full_title = re.sub(r'(\w)([A-Z])', r'\1 \2', full_title)
        full_title = re.sub(r'(\w)(-)', r'\1 \2', full_title)

        # Ensure correct handling of multi-element formulas such as "(CD3)4 NMn Cl3" -> "(CD3)4NMnCl3"
        full_title = re.sub(r'\b([A-Z][a-z]?)(\d*)\s*([A-Z][a-z]?)\s*([A-Z][a-z]?)(\d*)\b', r'\1\2\3\4\5', full_title)

        # Handle cases where "the" is joined with the next word
        full_title = re.sub(r'\bthe(\w)', r'the \1', full_title)

        # Replace multiple spaces with a single space
        full_title = normalize_whitespace(full_title)

        return full_title
    except Exception as e:
        print(f"Error extracting title: {str(e)}. Falling back to simple text extraction.")
        return title_elem.text.strip()

def download_pdf(title, authors, year, pdf_dir, arxiv_xml_folder):
    print(f"Attempting to retrieve paper: {title}")
    
    doi, abstract, scholar_url = get_paper_info(title, authors, year)
    
    # Try Sci-Hub with DOI - now with explicit parameter names
    filepath, pdf_url = try_scihub(doi=doi, output_dir=pdf_dir, authors=authors, year=year)
    if filepath:
        print(f"Found PDF via Sci-Hub at: {filepath}")
        return (os.path.basename(filepath), pdf_url, ''), doi
    
    # Try arXiv, with arxiv_xml_folder now passed in
    pdf_url = try_arxiv(title, arxiv_xml_folder)
    if pdf_url:
        print(f"Found PDF via arXiv: {pdf_url}")
        return download_file(pdf_url, title, authors, year, pdf_dir), doi
    
    print(f"Failed to retrieve PDF for '{title}'")
    return None, doi

last_crossref_request_time = 0

def get_paper_info(title, authors, year):
    global last_crossref_request_time
    
    doi, abstract, url = '', '', ''
    
    try:
        # Search for the paper using scholarly
        search_query = scholarly.search_pubs(title)
        paper = next(search_query)
        
        # Extract relevant information
        doi = paper.get('doi', '')
        abstract = paper.get('abstract', '')
        url = paper.get('url', '')
    except Exception as e:
        print(f"Error fetching paper info from scholarly: {str(e)}")
    
    # If DOI is not found, try Crossref
    if not doi:
        try:
            current_time = time.time()
            time_since_last_request = current_time - last_crossref_request_time
            
            if time_since_last_request < 61:
                wait_time = 61 - time_since_last_request
                print(f"Waiting {wait_time:.2f} seconds before making Crossref request...")
                time.sleep(wait_time)
            
            crossref_url = f"https://api.crossref.org/works?query.title={urllib.parse.quote(title)}&query.author={urllib.parse.quote(authors)}&filter=from-pub-date:{year},until-pub-date:{year}"
            response = requests.get(crossref_url)
            last_crossref_request_time = time.time()
            
            if response.status_code == 200:
                data = response.json()
                if data['message']['items']:
                    doi = data['message']['items'][0].get('DOI', '')
                    if not url:
                        url = data['message']['items'][0].get('URL', '')
        except Exception as e:
            print(f"Error fetching paper info from Crossref: {str(e)}")
    
    print(f"DOI found for '{title}': {doi}")
    return doi, abstract, url

def try_unpaywall(doi):
    if not doi:
        return None
    
    api_url = f"https://api.unpaywall.org/v2/{doi}?email=your_email@example.com"
    response = requests.get(api_url)
    
    if response.status_code == 200:
        data = response.json()
        best_oa_location = data.get('best_oa_location', {})
        if best_oa_location:
            return best_oa_location.get('url_for_pdf') or best_oa_location.get('url')
    
    return None

last_arxiv_request_time = 0
def try_arxiv(title, xml_dir):
    global last_arxiv_request_time
    
    current_time = time.time()
    time_since_last_request = current_time - last_arxiv_request_time
    
    wait_time = max(60, 60 - time_since_last_request)
    print(f"Waiting {wait_time:.2f} seconds before making arXiv request...")
    time.sleep(wait_time)
    
    arxiv_url = f"http://export.arxiv.org/api/query?search_query=ti:{title}&start=0&max_results=1"
    
    try:
        print(f"Querying arXiv for: {title}")
        response = requests.get(arxiv_url)
        last_arxiv_request_time = time.time()
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'xml')
            entry = soup.find('entry')
            if entry:
                # Save the XML content to a file
                file_name = f"{title.replace(' ', '_')}.xml"
                file_path = os.path.join(xml_dir, file_name)
                with open(file_path, 'w', encoding='utf-8') as xml_file:
                    xml_file.write(response.text)
                
                print(f"Saved arXiv XML for '{title}' to: {file_path}")
                
                # Continue extracting the PDF URL
                pdf_url = entry.find('link', {'title': 'pdf'})
                if pdf_url:
                    print(f"arXiv PDF found for: {title}")
                    return pdf_url['href']
        else:
            print(f"arXiv request failed with status code: {response.status_code}")
    except Exception as e:
        print(f"Error querying arXiv: {str(e)}")
    
    print(f"No arXiv PDF found for: {title}")
    return None

def normalize_author_name(authors_str):
    # Normalize author names from various formats into a consistent last name.
    # Handles multiple formats:
    if not authors_str or authors_str == "Unknown Authors":
        return "Unknown"
    
    print(f"Normalizing author string: {authors_str}")
    
    # Remove any parenthetical content and extra spaces
    authors_str = re.sub(r'\([^)]*\)', '', authors_str).strip()
    
    # If multiple authors, take only the first one
    if ',' in authors_str:
        # Handle "Sakai, T." format
        if authors_str.split(',')[1].strip().upper() == authors_str.split(',')[1].strip():
            # It's likely an initial after comma
            last_name = authors_str.split(',')[0].strip()
        else:
            # It might be separating multiple authors
            authors_str = authors_str.split(',')[0]
    
    # Split the name into parts
    parts = authors_str.split()
    if not parts:
        return "Unknown"
    
    # Remove periods from initials (e.g., "P." -> "P")
    parts = [p.rstrip('.') for p in parts]
    
    # If first parts are just single letters or initials (e.g., "T", "A.V."), skip them
    while parts and (len(parts[0]) == 1 or parts[0].replace('.', '').isupper()):
        if len(parts) > 1:  # Only remove if there are more parts left
            parts.pop(0)
        else:
            break
    
    if not parts:
        return "Unknown"
    
    # Take the last remaining part as the last name
    last_name = parts[-1]
    
    # Clean up any remaining punctuation
    last_name = re.sub(r'[^\w\s-]', '', last_name)
    
    print(f"Extracted last name: {last_name}")
    return last_name

def generate_pdf_filename(authors, year, title=None):
    # Generate consistent PDF filename from author and year
    author_name = normalize_author_name(authors)
    year_str = str(year) if year and year != "Unknown" else ""
    
    # If author is unknown and title is provided, use first word of title
    if author_name == "Unknown" and title:
        first_word = re.sub(r'[^\w\s-]', '', title.split()[0])
        author_name = f"Unknown_{first_word}"
    
    filename = f"{author_name}{year_str}.pdf"
    print(f"Generated filename: {filename} from authors: {authors}, year: {year}")
    return filename

def try_scihub(doi, output_dir='pdf_files', authors='Unknown', year='Unknown'):
    # Create a cloudscraper session to bypass DDoS protection
    scraper = cloudscraper.create_scraper(
        browser={
            'browser': 'chrome',
            'platform': 'windows',
            'mobile': False
        },
        delay=60  # Minimum delay of 60 seconds
    )
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
    }
    for domain in SCI_HUB_DOMAINS:
        try:
            # Construct the direct Sci-Hub URL
            url = f"{domain}{doi}"
            print(f"Trying to access: {url}")
            
            # Get the page with DDoS protection bypass
            response = scraper.get(url, headers=headers)
            
            if response.status_code == 200:
                # Parse the page to find the PDF link
                soup = BeautifulSoup(response.text, 'html.parser')
                pdf_iframe = soup.find('iframe', id='pdf')
                
                if pdf_iframe and pdf_iframe.get('src'):
                    pdf_url = pdf_iframe['src']
                    if not pdf_url.startswith('http'):
                        pdf_url = 'https:' + pdf_url if pdf_url.startswith('//') else domain + pdf_url
                    
                    print(f"Found PDF URL: {pdf_url}")
                    
                    # Add delay before downloading PDF
                    time.sleep(60)  # 60 second delay before PDF download
                    
                    # Download the PDF
                    pdf_response = scraper.get(pdf_url, headers=headers)
                    if pdf_response.status_code == 200 and 'application/pdf' in pdf_response.headers.get('Content-Type', ''):
                        # Generate filename using the new function
                        filename = generate_pdf_filename(authors, year)
                        filepath = os.path.join(output_dir, filename)
                        
                        # Save the PDF
                        with open(filepath, 'wb') as f:
                            f.write(pdf_response.content)
                        
                        print(f"Successfully downloaded PDF to: {filepath}")
                        return filepath, pdf_url

                    return None, None
                            
            print(f"No PDF found at {domain}, waiting 60 seconds before trying next domain...")
            time.sleep(60)  # 60 second delay between domain attempts
            
        except Exception as e:
            print(f"Error accessing {domain}: {str(e)}")
            print("Waiting 60 seconds before trying next domain...")
            time.sleep(60)  # 60 second delay after errors
            continue
    
    print(f"Failed to retrieve PDF for DOI: {doi} from any domain")
    return False

# Helper function to initialize cloudscraper with proper settings
def create_scraper_session():
    return cloudscraper.create_scraper(
        browser={
            'browser': 'chrome',
            'platform': 'windows',
            'mobile': False
        }
    )

def download_file(url, title, authors, year, pdf_dir):
    try:
        print(f"Attempting to download file for: {title}")
        print(f"Author information: {authors}")
        print(f"Year: {year}")
        
        response = requests.get(url, stream=True)
        response.raise_for_status()
        
        content_type = response.headers.get('Content-Type', '').lower()
        if 'application/pdf' not in content_type:
            print(f"Warning: Content type is not PDF: {content_type}")
        
        # Generate filename using the new function
        file_name = generate_pdf_filename(authors, year, title)
        print(f"Generated filename: {file_name}")
        
        # Use search session subfolder
        file_path = os.path.join(pdf_dir, file_name)
        
        # Check if file already exists and handle duplicates
        counter = 1
        original_name = file_name.replace('.pdf', '')
        while os.path.exists(file_path):
            file_name = f"{original_name}_{counter}.pdf"
            file_path = os.path.join(pdf_dir, file_name)
            counter += 1
        
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        
        if os.path.getsize(file_path) > 0:
            print(f"Successfully downloaded and saved PDF: {file_name}")
            return file_name, url, ''
        else:
            print(f"Downloaded file is empty: {file_name}")
            os.remove(file_path)
            return "Download failed", "", ""
            
    except Exception as e:
        print(f"Error downloading file: {str(e)}")
        traceback.print_exc()  # Add stack trace for better debugging
        return "Download failed", "", ""

def save_to_excel(data, excel_path):
    wb = Workbook()
    ws = wb.active
    
    # Base headers that are always included
    base_headers = ['No.', 'Title', 'Authors', 'Year', 'Keywords', 'Google Scholar Link', 'DOI', 'PDF Filename', 'PDF Download Link', 'First 100 Sentences']
    
    # Insert optional headers if the data contains these fields
    headers = base_headers.copy()
    if any('Compound' in paper for paper in data):
        headers.insert(1, 'Compound')
    if any('Spin' in paper for paper in data):
        headers.insert(2 if 'Compound' in headers else 1, 'Spin')

    # Define styles (keeping original styling)
    header_style = Font(bold=True, size=11)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Write headers with styling (keeping original styling)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_style
        cell.alignment = center_aligned
        # Add light gray background to headers
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Write data with specified formatting 
    for row, paper in enumerate(data, 2):
        # No. column
        cell = ws.cell(row=row, column=1, value=row-1)
        cell.alignment = center_aligned

        # Other columns
        for col, key in enumerate(headers[1:], 2):
            value = paper.get(key, "")
            if isinstance(value, str):
                value = ''.join(char for char in value if char.isprintable())
            
            cell = ws.cell(row=row, column=col, value=value)
            
            # Apply specific formatting based on column type
            if key in ['No.', 'Year', 'Compound', 'Spin']:
                cell.alignment = center_aligned
            else:
                cell.alignment = left_aligned

    # Set optimized column widths 
    fixed_widths = {
        'A': 5,    # No.
        'B': 15,   # Compound/Title
        'C': 15,   # Spin/Authors
        'D': 75,   # Title/Year
        'E': 45,   # Authors/Keywords
        'F': 6,    # Year/Google Scholar Link
        'G': 55,   # Keywords/DOI
        'H': 35,   # Google Scholar Link/PDF Filename
        'I': 25,   # DOI/PDF Download Link
        'J': 20,   # PDF Filename/First 100 Sentences
        'K': 35,   # PDF Download Link
        'L': 40    # First 100 Sentences
    }

    # Apply fixed column widths
    for col_letter, width in fixed_widths.items():
        if col_letter <= chr(64 + len(headers)):  # Only set width for columns that exist
            ws.column_dimensions[col_letter].width = width

    # Add freeze panes to keep headers visible
    ws.freeze_panes = 'A2'

    # Add filters to headers (updated to use dynamic column count)
    last_col = chr(64 + len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(data) + 1}"

    # Set row height for content rows
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 25  # Slightly increased height for better readability

    try:
        wb.save(excel_path)
        print(f"Excel file saved successfully at {excel_path}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        return False

def extract_text_from_pdf(pdf_path):
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            
            # Clean up the text
            text = re.sub(r'\s+', ' ', text)  # Replace multiple spaces with single space
            text = text.strip()
            
            if not text:
                raise ValueError("No text extracted from PDF")
            
            return text
    except Exception as e:
        print(f"Error extracting text from PDF: {str(e)}")
        return ""

def extract_first_100_sentences(pdf_path):
    try:
        text = extract_text_from_pdf(pdf_path)
        if not text:
            return "Failed to extract text from PDF."
        
        print(f"Extracted text length: {len(text)}")
        
        # Simple sentence splitting
        sentences = re.split(r'(?<=[.!?])\s+', text)
        result = ' '.join(sentences[:100])
        
        if not result:
            return "Failed to extract meaningful sentences from PDF."
        
        # Print only the first 10 sentences in the command prompt
        print(f"First 10 sentences from PDF '{pdf_path}':")
        for i, sentence in enumerate(sentences[:10], start=1):  # Limiting to 10
            print(f"{i}: {sentence}")

        print(f"Extracted {len(sentences[:100])} sentences")
        return result
    except Exception as e:
        print(f"Error extracting sentences from PDF {pdf_path}: {str(e)}")
        return f"Error extracting sentences: {str(e)}"

def is_theoretical_paper(text):
    """Helper function to detect theoretical papers based on content"""
    if not text:
        return False
        
    theoretical_indicators = [
        ('hamiltonian', 3),
        ('spin chain', 2),
        ('theoretical', 2),
        ('model', 2),
        ('exact diagonalization', 3),
        ('quantum', 2),
        ('classical', 2),
        ('numerical', 1)
    ]
    
    # Convert to lowercase for case-insensitive matching
    text_lower = text[:1000].lower()  # Check first 1000 characters only
    
    # Count weighted occurrences of theoretical terms
    theory_score = sum(weight for term, weight in theoretical_indicators 
                      if term in text_lower)
    
    # If we have a significant theory score in the first part of the text
    return theory_score >= 5

def is_valid_compound_name(compound, text_context):
    # Common scientific abbreviations to exclude
    exclude_terms = {
        'PDF', 'USA', 'IOP', 'DOI', 'URL', 'XML', 'API', 'FIG', 'TABLE', 
        'DMI', 'AFM', 'FM', 'BEC', 'VOL', 'DNA', 'INTRODUCTION', 'ABSTRACT'
    }
    
    if compound in exclude_terms:
        return False
        
    # Check if the compound appears in a materials-related context
    context_keywords = [
        'compound', 'material', 'crystal', 'sample', 'synthesized',
        'grown', 'prepared', 'measured', 'doped'
    ]
    
    text_context = text_context.lower()
    
    # Look for pattern indicating chemical formula explanation
    formula_pattern = rf"{compound}\s*\([A-Za-z0-9\s,]+\)"
    has_formula = bool(re.search(formula_pattern, text_context))
    
    # Check for material-related context
    has_context = any(keyword in text_context for keyword in context_keywords)
    
    return has_formula or has_context

def extract_compound_from_title_or_pdf(title, pdf_filename, pdf_dir):
    # First check if this is a theoretical paper
    theoretical_keywords = [
        'heisenberg', 'classical', 'quantum', 'theory', 'model', 'hamiltonian',
        'theoretical', 'xy chain', 'ising', 'spin chain'
    ]
    
    # Convert title to lowercase for case-insensitive matching
    title_lower = title.lower()
    
    # Check if multiple theoretical keywords are present
    theory_count = sum(1 for keyword in theoretical_keywords if keyword in title_lower)
    
    if theory_count >= 2:
        return "Theoretical model"
        
    # If not theoretical, proceed with original compound extraction
    print(f"Attempting to extract compound from the title: {title}")
    compound_from_title = extract_compound_name(title)
    
    if compound_from_title and compound_from_title != "Unknown" and compound_from_title != "No compound information found from abstract":
        # Additional validation for chemical formulas
        if re.match(r'^[A-Z][a-z]?\d*(?:[A-Z][a-z]?\d*)*$', compound_from_title) or \
           '(' in compound_from_title or ')' in compound_from_title:  # For complex formulas like (CD3)4NMnCl3
            print(f"Successfully extracted compound from the title: {compound_from_title}")
            return compound_from_title
        else:
            # Check if we have any fragments or abbreviations
            if len(compound_from_title) <= 4 and compound_from_title.isupper():  # Like "CO" or "AFM"
                context = extract_text_from_pdf(os.path.join(pdf_dir, pdf_filename)) if pdf_filename else ""
                if is_theoretical_paper(context):
                    return "Theoretical model"

    # Enhanced PDF extraction logic
    if pdf_filename and pdf_filename != "Download failed":
        print(f"Attempting to extract compound from the PDF: {pdf_filename}")
        try:
            pdf_path = os.path.join(pdf_dir, pdf_filename)
            full_text = extract_text_from_pdf(pdf_path)
            if not full_text:
                return "Failed to extract text from PDF"

            text_lower = full_text.lower()

            # Reference compound patterns
            reference_patterns = [
                r'compared (?:to|with) ([^.]*?(?:[A-Z][a-z0-9]+(?:\d+)?)+[^.]*)',
                r'similar to ([^.]*?(?:[A-Z][a-z0-9]+(?:\d+)?)+[^.]*)',
                r'such as ([^.]*?(?:[A-Z][a-z0-9]+(?:\d+)?)+[^.]*)',
                r'properties of ([^.]*?(?:[A-Z][a-z0-9]+(?:\d+)?)+[^.]*)',
                r'like (?:the|those of) ([^.]*?(?:[A-Z][a-z0-9]+(?:\d+)?)+[^.]*)'
            ]

            # First check for reference compounds
            referenced_compounds = []
            for pattern in reference_patterns:
                matches = re.finditer(pattern, full_text)
                for match in matches:
                    context = match.group(1)
                    # Look for chemical formulas in the matched context
                    formulas = re.findall(r'\b(?:[A-Z][a-z]?\d*)+\b', context)
                    for formula in formulas:
                        context_start = max(0, match.start() - 150)
                        context_end = min(len(full_text), match.end() + 150)
                        surrounding_context = full_text[context_start:context_end]
                        if formula not in referenced_compounds:
                            referenced_compounds.append({
                                'compound': formula,
                                'context': surrounding_context,
                                'score': 9  # High score for referenced compounds
                            })

            # Check if paper is theoretical
            theoretical_indicators = [
                'hamiltonian', 'quantum field', 'theoretical model', 'equation of motion',
                'mathematical model', 'theoretical framework', 'analytical solution',
                'numerical simulation', 'quantum theory', 'theoretical study'
            ]
            
            experimental_indicators = [
                'measured', 'synthesized', 'prepared', 'experiment', 
                'sample', 'crystal growth', 'characterized', 'observation'
            ]
            
            theory_count = sum(1 for term in theoretical_indicators if term in text_lower)
            experimental_count = sum(1 for term in experimental_indicators if term in text_lower)
            
            # Special handling for theoretical papers
            is_theoretical = theory_count > experimental_count and theory_count > 3
            if is_theoretical:
                # If we found referenced compounds in a theoretical paper, return the best one
                if referenced_compounds:
                    return referenced_compounds[0]['compound']
                # If purely theoretical with no compounds mentioned, return theoretical paper
                if experimental_count == 0:
                    return "Theoretical paper - no physical compound"

            # Common scientific terms/abbreviations to exclude
            exclude_terms = {
                'PDF', 'USA', 'IOP', 'DOI', 'URL', 'XML', 'API', 'FIG', 'TABLE', 
                'NATURE', 'CONCLUSION', 'INTRODUCTION', 'ABSTRACT', 'REF', 'IEEE',
                'DMI', 'AFM', 'FM', 'BEC', 'VOL', 'DNA',
                'JETP', 'PHYSICAL', 'REVIEW', 'LETTERS', 'SCIENCE', 'ACS', 'RSC',
                'SPINCORRELATIO', 'MAGNETIZATION', 'HAMILTONIAN', 'CORRELATION',
                'SUPPLEMENTARY', 'APPENDIX', 'CHAPTER', 'ARXIV', 'THESIS',
                'JOURNAL', 'PHYSICS', 'CHEMISTRY', 'MATERIALS',
                'EQUATION', 'DISCUSSION', 'THEORY', 'METHOD', 'RESULTS',
                'SISSA', 'MECHANICS', 'MODELING', 'SIMULATION', 'CALCULATION'
            }
            
            compounds = referenced_compounds.copy()  # Start with any referenced compounds

            # Look for compounds in regular context
            # First look for abbreviated compounds with their formulas
            abbrev_formula_patterns = [
                r'\b([A-Z]{3,})\s*\(([A-Z][A-Za-z0-9\(\)]+)\)',  # TMMC (MnCl3)
                r'\b([A-Z]{3,})\s*\(.*?([A-Z][A-Za-z0-9\(\)]+)\)',  # TMMC (full name with formula)
            ]
            
            for pattern in abbrev_formula_patterns:
                matches = re.finditer(pattern, full_text)
                for match in matches:
                    abbrev = match.group(1)
                    formula = match.group(2)
                    if abbrev.upper() not in exclude_terms:
                        context_start = max(0, match.start() - 150)
                        context_end = min(len(full_text), match.end() + 150)
                        context = full_text[context_start:context_end].lower()
                        
                        if any(indicator in context for indicator in [
                            'compound', 'material', 'crystal', 'sample', 'magnetic'
                        ]):
                            compounds.append({
                                'compound': f"{abbrev} ({formula})",
                                'context': context,
                                'score': 8,
                                'source': 'abbreviation_with_formula'
                            })
            
            # Look for complex molecular formulas
            complex_matches = re.finditer(r'[A-Z][a-z]?\([A-Z0-9a-z]+\)\d*(?:[A-Z][a-z]?\d*)*', full_text)
            for match in complex_matches:
                formula = match.group()
                context_start = max(0, match.start() - 150)
                context_end = min(len(full_text), match.end() + 150)
                context = full_text[context_start:context_end].lower()
                
                if any(indicator in context for indicator in [
                    'compound', 'crystal', 'sample', 'synthesized', 'structure'
                ]):
                    compounds.append({
                        'compound': formula,
                        'context': context,
                        'score': 7,
                        'source': 'complex_formula'
                    })

            # Look for chemical formulas near specific keywords
            formula_patterns = [
                (r'(?:[A-Z][a-z]?\d*)+[A-Z][a-z]?\d*O\d*', 6),  # Oxide compounds
                (r'[A-Z][a-z]?\d*(?:[A-Z][a-z]?\d*)+', 5),      # Multi-element compounds
                (r'[A-Z][a-z]?(?:\([IVX]+\))', 4)               # Compounds with oxidation states
            ]
            
            context_keywords = [
                'antiferromagnet', 'ferromagnet', 'magnetic', 'crystal',
                'compound', 'material', 'sample', 'structure', 'synthesized',
                'measured', 'doped', 'grown'
            ]
            
            for pattern, base_score in formula_patterns:
                formula_matches = re.finditer(pattern, full_text)
                for match in formula_matches:
                    formula = match.group()
                    if not re.search(r'[A-Z][a-z]?\d*[A-Z]', formula):
                        continue
                        
                    context_start = max(0, match.start() - 150)
                    context_end = min(len(full_text), match.end() + 150)
                    context = full_text[context_start:context_end].lower()
                    
                    if formula.upper() not in exclude_terms:
                        score = base_score
                        for keyword in context_keywords:
                            if keyword in context:
                                score += 1
                                
                        compounds.append({
                            'compound': formula,
                            'context': context,
                            'score': score,
                            'source': 'chemical_formula'
                        })

            if compounds:
                # Sort by score and complexity
                compounds.sort(key=lambda x: (
                    x['score'],
                    len(re.findall(r'[A-Z]', x['compound'])),  # Number of elements
                    len(x['compound'])  # Total length as tiebreaker
                ), reverse=True)
                
                print(f"Found compounds (showing top 3):")
                for compound in compounds[:3]:
                    print(f"  {compound['compound']} (score: {compound['score']}, source: {compound.get('source', 'reference')})")
                
                return compounds[0]['compound']

            return "No valid compound found in PDF"

        except Exception as e:
            print(f"Error extracting compound from PDF: {str(e)}")
            return "Error occurred while extracting compound"

    print("No compound found in both title and PDF.")
    return "No compound found"

def analyze_saved_html(html_parsing_folder):
    for filename in os.listdir(html_parsing_folder):
        if filename.endswith('_raw.html'):
            file_path = os.path.join(html_parsing_folder, filename)
            with open(file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            soup = BeautifulSoup(html_content, 'html.parser')
            title_elem = soup.select_one('.gs_rt')
            
            print(f"Analyzing {filename}:")
            print("Raw title element:")
            print(title_elem)
            print("\nExtracted title:")
            print(extract_title(title_elem))
            print("\n" + "="*50 + "\n")

def read_max_results_from_file(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        for i, line in enumerate(lines):
            if "Specify the maximum number of searches" in line:
                # Read the next line, which should contain the number
                try:
                    return int(lines[i + 1].strip())
                except ValueError:
                    raise ValueError("The value for 'max_results' in inputs.txt is missing or invalid.")

def is_valid_quantum_spin(numerator, denominator=None):
    # Validates if a given spin value follows quantum mechanical principles.
    # Rules:
    # 1. Spin must be non-negative
    # 2. Must be either integer (0, 1, 2) or half-integer (1/2, 3/2, 5/2)
    # 3. Elementary particles rarely have spin > 2 in the Standard Model
    # 4. Denominator, if present, must be 2 (for half-integer spins)
    try:
        if denominator:
            # For fraction form (half-integer spins)
            if denominator != 2:  # Only /2 is allowed for half-integer spins
                return False
                
            spin_value = float(numerator) / float(denominator)
            # Must be a positive half-integer (1/2, 3/2, 5/2, etc.)
            return spin_value > 0 and (numerator % 2) == 1 and spin_value <= 2
        else:
            # For integer spins
            spin_value = int(numerator)
            # Must be a non-negative integer (0, 1, 2)
            return 0 <= spin_value <= 2
    except (ValueError, TypeError):
        return False

def extract_spin_value(text):
    
    # Extract physically valid spin values following quantum mechanical principles:
    # - Integer spins: S=0, S=1, S=2
    # - Half-integer spins: S=1/2, S=3/2, S=5/2
    
    # Also accepts:
    # - Forms with spaces: S = 1/2, S = 1
    # - Spin/spin notation: Spin-1/2, spin - 1/2
    # - Various spacing patterns: Spin - 1 / 2, spin-1/2
    # - Joined words: InteractingSpinSystem-1/2
    # First normalize spacing around 'Spin' keyword when it's joined with other words
    text = re.sub(r'([a-z])([Ss]pin)', r'\1 \2', text, flags=re.IGNORECASE)
    text = re.sub(r'([Ss]pin)([a-zA-Z])', r'\1 \2', text, flags=re.IGNORECASE)
    
    patterns = [
        # S= format with optional spaces
        r'S\s*=\s*(\d+)\s*(?:/\s*(\d+))?(?![0-9/])',
        # Spin- format with optional spaces, case insensitive, allowing for joined words
        r'(?:^|\s|[^a-zA-Z])[Ss]pin\s*-\s*(\d+)\s*(?:/\s*(\d+))?(?![0-9/])',
        # Spin= format (less common but possible)
        r'(?:^|\s|[^a-zA-Z])[Ss]pin\s*=\s*(\d+)\s*(?:/\s*(\d+))?(?![0-9/])',
        # Handle cases where numbers are separated, like "Spin - 1 /2 and 1"
        r'(?:^|\s|[^a-zA-Z])[Ss]pin\s*-\s*(\d+)\s*(?:/\s*(\d+))?\s*(?:and|,|&)?\s*(\d+)?(?![0-9/])'
    ]
    
    # First try to extract values from each pattern
    for pattern in patterns:
        matches = re.finditer(pattern, text)
        for match in matches:
            numerator = match.group(1)
            denominator = match.group(2) if len(match.groups()) > 1 else None
            
            # Remove any remaining spaces in the numbers
            numerator = numerator.strip()
            if denominator:
                denominator = denominator.strip()
                
            try:
                if denominator:
                    if not is_valid_quantum_spin(int(numerator), int(denominator)):
                        continue
                    return f"S={numerator}/{denominator}"
                else:
                    if not is_valid_quantum_spin(int(numerator)):
                        continue
                    return f"S={numerator}"
            except ValueError:
                continue

            # Check for additional integer spin value in the same match
            if len(match.groups()) > 2 and match.group(3):
                additional_spin = match.group(3).strip()
                try:
                    if is_valid_quantum_spin(int(additional_spin)):
                        # Return both spin values
                        if denominator:
                            return f"S={numerator}/{denominator} and S={additional_spin}"
                        else:
                            return f"S={numerator} and S={additional_spin}"
                except ValueError:
                    pass
    
    return "No valid quantum spin value found"

def extract_spin_from_title_or_pdf(title, pdf_filename, pdf_dir):
    # Attempt to extract spin value from title or PDF content.
    print(f"Attempting to extract spin value from the title: {title}")
    spin_from_title = extract_spin_value(title)
    
    if spin_from_title != "No valid quantum spin value found":
        print(f"Successfully extracted spin value from the title: {spin_from_title}")
        return spin_from_title

    # Check PDF if title doesn't contain spin value
    if pdf_filename and pdf_filename != "Download failed":
        print(f"Attempting to extract spin value from the PDF: {pdf_filename}")
        try:
            pdf_path = os.path.join(pdf_dir, pdf_filename)
            full_text = extract_text_from_pdf(pdf_path)
            if not full_text:
                return "Failed to extract text from PDF"

            # First try the abstract or first few paragraphs
            initial_text = full_text[:2000]  # Check first 2000 characters
            spin_value = extract_spin_value(initial_text)
            if spin_value != "No valid quantum spin value found":
                return spin_value

            # If not found in the beginning, check the entire text
            spin_value = extract_spin_value(full_text)
            if spin_value != "No valid quantum spin value found":
                return spin_value

            return "No valid quantum spin value found in PDF"

        except Exception as e:
            print(f"Error extracting spin value from PDF: {str(e)}")
            return "Error occurred while extracting spin value"

    print("No valid quantum spin value found in title or PDF.")
    return "No valid quantum spin value found"

def main():
    print(AUTHOR_INFO)
    
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    inputs_file = os.path.join(root_dir, 'inputs.txt')
    
    inputs = read_inputs(inputs_file)
    print("Inputs read:", inputs)  # Debug print
    
    search_name = inputs.get('1', '')
    max_results = read_max_results_from_file(inputs_file)
    keywords = inputs.get('3', '')
    search_compound = inputs.get('4', '').strip().upper() == 'Y'  # New: Get compound search preference
    search_spin = inputs.get('5', '').strip().upper() == 'Y'      # New: Get spin search preference
    
    print(f"Search name: {search_name}")
    print(f"Max results: {max_results}")
    print(f"Keywords: {keywords}")
    print(f"Search for compounds: {search_compound}")  # debug print
    print(f"Search for spin values: {search_spin}")    # debug print
    
    if not search_name or not keywords:
        print("Error: Search name or keywords are missing. Please check your inputs.txt file.")
        return
    
    pdf_dir, excel_dir, sentences_dir = create_directories(root_dir, search_name)

    arxiv_xml_folder = os.path.join(root_dir, 'arXiv_xml', search_name)
    os.makedirs(arxiv_xml_folder, exist_ok=True)
    
    html_parsing_folder = os.path.join(root_dir, 'html_parsing', search_name)
    os.makedirs(html_parsing_folder, exist_ok=True)
    
    try:
        data = search_papers(keywords, max_results, search_name, html_parsing_folder)
        
        if not data:
            print("No data collected from the search.")
        else:
            print(f"Collected data for {len(data)} papers")

            analyze_saved_html(html_parsing_folder)
                    
            for paper in data:
                try:
                    (pdf_filename, pdf_link, _), doi = download_pdf(paper['Title'], paper['Authors'], paper['Year'], pdf_dir, arxiv_xml_folder)
                    paper['PDF Filename'] = pdf_filename
                    paper['PDF Download Link'] = pdf_link
                    paper['DOI'] = doi

                    # Only search for compound if enabled
                    if search_compound:
                        compound = extract_compound_from_title_or_pdf(paper['Title'], pdf_filename, pdf_dir)
                        paper['Compound'] = compound
                    
                    # Only search for spin if enabled
                    if search_spin:
                        spin_value = extract_spin_from_title_or_pdf(paper['Title'], pdf_filename, pdf_dir)
                        paper['Spin'] = spin_value

                    if pdf_filename != "Download failed":
                        pdf_path = os.path.join(pdf_dir, pdf_filename)
                        first_100_sentences = extract_first_100_sentences(pdf_path)
                        if first_100_sentences:
                            paper['First 100 Sentences'] = first_100_sentences
                            save_first_100_sentences(first_100_sentences, pdf_filename, sentences_dir)
                        else:
                            paper['First 100 Sentences'] = "Failed to extract sentences from PDF."
                    else:
                        paper['First 100 Sentences'] = "PDF download failed. No sentences extracted."

                except Exception as e:
                    print(f"Failed to process paper '{paper['Title']}': {str(e)}")
                    paper['PDF Filename'] = "Processing failed"
                    paper['PDF Download Link'] = ""
                    paper['DOI'] = ""
                    paper['First 100 Sentences'] = f"Error occurred: {str(e)}"
                
                time.sleep(max(60, 60 + (10 * hash(paper['Title']) % 11)))

            # Prepare data for saving
            excel_path = os.path.join(excel_dir, f"{search_name}.xlsx")
            
            # Remove Compound and Spin from data if not requested
            if not search_compound:
                for paper in data:
                    if 'Compound' in paper:
                        del paper['Compound']
            
            if not search_spin:
                for paper in data:
                    if 'Spin' in paper:
                        del paper['Spin']
            
            # Call save_to_excel with just data and path
            if save_to_excel(data, excel_path):
                print(f"Excel file created at {excel_path}")

            else:
                print("Failed to create Excel file. Attempting to save as CSV...")
                csv_path = os.path.join(excel_dir, f"{search_name}.csv")
                try:
                    headers = ['No.', 'Compound', 'Title', 'Authors', 'Year', 'Keywords', 'Google Scholar Link', 'DOI', 'PDF Filename', 'PDF Download Link', 'First 100 Sentences']
                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=headers)
                        writer.writeheader()
                        for i, paper in enumerate(data, 1):
                            row = {'No.': i}
                            row.update({k: str(v)[:32767] if isinstance(v, str) else v for k, v in paper.items()})
                            writer.writerow(row)
                    print(f"CSV file created at {csv_path}")
                except Exception as e:
                    print(f"Failed to create CSV file: {str(e)}")

    except Exception as e:
        print(f"An error occurred during execution: {str(e)}")

if __name__ == "__main__":
    main()